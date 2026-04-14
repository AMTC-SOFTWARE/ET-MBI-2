from copy import copy
import requests
import openpyxl
import json
import os
from model import model
import gc

datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2,printerhost=datos_conexion.datos_acceso()


fuses_color = {
    #"1":    "negro", HMTEST ILX296270B1031517 EL.
    "5A"  :   {"N000000008698":"beige", "N000000008708":"beige", "N000000004202":"beigeClear", "N000000006465":"beige"},
    "7.5A":   {"N000000008699":"cafe", "N000000008709":"cafe", "N000000006466":"cafe"},
    "10A" :   {"N000000008700":"rojo", "N000000008710":"rojo", "N000000004204":"rojoClear"},
    "15A" :   {"N000000008701":"azul", "N000000008711":"azul"},
    "20A" :   {"N000000008702":"amarillo"},
    "25A" :   {"N000000008703":"natural"},
    "30A" :   {"N000000008704": "verde", "N000000007658":"verde"},
    "40A" :   {"N000000007659": "naranja"},
    "50A" :   {"N000000007660":"rojo"},
    "60A" :   {"A0009821923":"1008695"},
    "70A" :   {"A0025429419":"1010733"}
    # "60":   "azul"
    }

def modulesConfig(data):
    pieces = stagingModules(data)
    config(pieces)


def config(data):
    print("Configurando Modulos")
    #print("TABLAAAAA Vision: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{data}/modulos_configuracion/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    

def titlesCol(sheet,subensamble,cavidad,descripcion,mercedez,amp, startmodule): 
    
    for cell in sheet[4]:
        
            
            if 'ubensamble' in str(cell.value) :
                subensamble = cell.column
            elif 'kurzname' in str(cell.value) :
                cavidad = cell.column
            elif 'escripcion' in str(cell.value) :
                descripcion = cell.column
            elif 'ercedez' in str(cell.value) :
                mercedez = cell.column
            elif 'Amp' in str(cell.value) or 'amp' in str(cell.value) :
                amp = cell.column

            if amp and subensamble and cavidad and descripcion and mercedez:
                break


    for cell in sheet[3]:
        if 'A' in str(cell.value) :
            startmodule = cell.column
            break

    return subensamble, cavidad, descripcion, mercedez, amp, startmodule


def stagingModules(data):
    print("Preparando Modulos", data)
    pieces = {}
    file_name = None
    dir_path = os.path.join(os.getcwd(), '..\\modules\\')

    ## Revisamos los modulos de configuracion 
    colega = f"http://{host}:5000/api/get/{data}/modulos_configuracion_staging/all/-/-/-/-/-"
    existing = requests.get(colega).json()

    
    print("Existing: ", existing)
    if existing['items'] == 0:
        
    

    
        print("No hay modulos de configurados en la tabla original, se procede a configurar")
        
        for root, dirs, files in os.walk(dir_path):
            for file_name in files: 
                 if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                    
                    file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                    sheets = file.worksheets

                    modules_data = {}

                    for sheet in sheets:
                        print("Leyendo hoja: ", sheet)
                        if sheet.sheet_state != "visible":
                            continue
                     
                        #print("Sheet: ", sheet)
                        currentSheet = file[sheet.title]
                        ## Hay que encontrar lo que hay  y por donde empieza
                        ## Se garantiza que el documento siempre tenga en la fila 3 el inicio de los modulos
                        ## En la fila 4 se garantiza que se encuentren los titulos de cada columna, por lo tanto se pueden usar para identificar cada propiedad de la posicion
                        subensamble = None
                        cavidad = None
                        descripcion = None
                        mercedez = None
                        amp = None
                        startmodule = None

                        # titlesCol devuelve una tupla con las columnas detectadas;
                        # hay que asignar esos valores aquí para conservarlos.
                        subensamble, cavidad, descripcion, mercedez, amp, startmodule = \
                            titlesCol(currentSheet, subensamble, cavidad, descripcion, mercedez, amp, startmodule)

                        # print(
                        #       "subensamble: ", subensamble,
                        #       "cavidad: ", cavidad,
                        #       "descripcion: ", descripcion,
                        #       "mercedez: ", mercedez,
                        #       "amp: ", amp,
                        #       "startmodule: ", startmodule
                        #       )
                         

                        
                        
                        # title = currentSheet.cell(row=5, column=descripcion).value.lower().strip() +"s" # se obtiene el titulo de la columna de descripcion, se convierte a minuscula y se pluraliza para usarlo como llave del diccionario
                        # if title == 'relays':
                        #     title = 'fusibles'
                        
                        posicion = currentSheet.cell(row=5, column=subensamble).value.upper() # se obtiene el titulo de la columna de subensamble, se convierte a mayuscula para usarlo como valor de posicion del diccionario

                        

                       
                        

                        for column in range(startmodule, currentSheet.max_column + 1):
                            module = currentSheet.cell(row = 3, column = column).value #se obtiene el valor de la celda que contiene el nombre del módulo

                            if module not in modules_data:

                                modules_data[module] = {
                                    # title: {"cajas": []}
                                }
                            # if title not in modules_data[module]:
                            #     modules_data[module][title] = {"cajas": []}



                            if isinstance(module,str):
                                module = module.replace(" ","")#se eliminan posibles espacios existentes, solo en str, porque puede haber valores None
                            print("Modulo: ",module, "Config")

                            

                            dato = {
                                     "posicion": f"{posicion}",
                                     #f"{title}": [ ]
                                     }
                            for row in range(5,currentSheet.max_row  + 1):
                                value = currentSheet.cell(row = row, column = column).value #se obtiene el valor de la celda ej: "X"
                                amperaje = ""
                                propiedad = ""
                                describe = ""
                                if isinstance(value,str):
                                    value = value.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                if value == "x" or value == "X":
                                    
                                    # Convertir de forma segura las lecturas de celda a string
                                    _box = currentSheet.cell(row = row, column = subensamble).value
                                    box = "" if _box is None else str(_box).strip()

                                    _describe = currentSheet.cell(row = row, column = descripcion).value
                                    describe = "" if _describe is None else str(_describe).strip()

                                    if describe == "Multifuse" or describe == "Relay":
                                        describe = "Fusible"

                                    if describe not in modules_data[module]:
                                        modules_data[module][describe] = {"cajas": []}


                                    if describe not in dato:
                                        dato[describe] = []


                                    
                                    if mercedez is not None:
                                        _numMercedez = currentSheet.cell(row = row, column = mercedez).value
                                        print("Mercedes raw value: ", _numMercedez)
                                        numMercedez = "" if _numMercedez is None else str(_numMercedez).strip()
                                    
                                    if amp is not None:
                                        _amperaje = currentSheet.cell(row = row, column = amp).value
                                        amperaje = "" if _amperaje is None else str(_amperaje).strip()
                                        propiedad = fuses_color[amperaje][numMercedez] #Obteniendo el valor real de color dependiendo del amperaje y el numero de mercedez, usando ambos como llaves del diccionario
                                    else:
                                        #Sabemos que es valido tener la cavidad pero no tiene un aspecto fisico detectable como el color, por lo tanto se asigna un valor generico que indica que la propiedad existe pero no tiene un valor fisico definido
                                        propiedad = "true"

                                    _cavidad = currentSheet.cell(row = row, column = cavidad).value
                                    cavidadValue = "" if _cavidad is None else str(_cavidad).strip()

                                    if isinstance(box,str):
                                        box = box.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None
                                    # print({
                                    #     "CAJA": box,
                                    #     "CAVIDAD": cavidad,
                                    #     "DESCRIPCION": descripcion,
                                    #     "MERCEDES": mercedez,
                                    #     "AMP": amp
                                    # })
                                    
                                    
                                    dato[describe].append(
                                        {"zona": f"{cavidadValue}", "propiedad": f"{propiedad}"}
                                    )


                            clave = list(dato.keys())[1] if len(dato.keys()) > 1 else None # se obtiene la clave del diccionario dato que no sea "posicion", es decir, la descripcion del componente, para usarla como llave en el diccionario modules_data
                            if clave:

                                modules_data[module][clave]["cajas"].append(dato)

                            else:
                                modules_data[module]["vacio"] = {"cajas": []}
                                
                                #modules_data[module]["vacio"]["cajas"].append({})
                            #print("El DATO: ", dato)
                    print("Modules data: ", modules_data)
                    modePosting(modules_data, data)
    #Transferir datos a tabla principal, esto se hace al final para evitar hacer transferencias parciales en caso de que el proceso de lectura y posteo de los módulos tarde mucho y se quiera evitar que la tabla principal tenga datos incompletos durante ese tiempo                
    transfer_endpoint = f"http://{host}:5000/api/archive-and-empty/modulos_configuracion_staging/{data}"
    transfer_response = requests.post(transfer_endpoint)                


def modePosting(modules_data, data):
    '''Ya que tenemos toda la informacion organizada,tenemos que filtar los valores vacios e irrelevantes,
        la declaracion de modulos sin importar que esten vacios deben ser declaraos al menos una vez para declarar que fueron recibidos '''
    print("Enviando modulos a la base de datos")

    endpoint = f"http://{host}:5000/api/post/modulos_configuracion_staging"


    

    for module in modules_data:
        for describe in modules_data[module]:
            #print("Module: ", module)
            #print("Describe: ", describe)

            
            # primero busquemos que existe en el modulo que se lee
            cajas = modules_data[module][describe]['cajas']
            #print(cajas)

            query = {
                "DBEVENT": data,
                "MODULO": module,
                "TIPO": describe,    
            }
            if describe != "vacio":
                for i in range(len(cajas)):
                    caja = cajas[i]
                    numeroCaja = i + 1
                    
                    query[f"CAJA_{numeroCaja}"] = caja

                    # print("Numero de caja: ", numeroCaja)
                #print(query)
            

            '''INSERT INTO `evento_ueueueue_x296_izquierda`.`modulos_configuracion_staging` (`ID`, `MODULO`, `TIPO`, `CAJA_1`, `CAJA_2`) VALUES ('1', 'A2965403846', 'Fusible', '{\'posicion\': \'PDC-D\', \'Fusible\': [{...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}, {...}]}', '{\'posicion\': \'PDC-P\', \'Fusible\': [{...}, {...}], \'Multifuse\': [{...}]}');   '''

            
            response = requests.post(endpoint, json=query)
            response_data = response.json()
            #print("Response data:", response_data)

    
            




                        

                                        
    